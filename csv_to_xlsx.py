import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Cargar el CSV en un DataFrame
#csv_file = "test.csv"  # Nombre del archivo CSV
csv_file = "stations.csv"  # Nombre del archivo CSV

# Intentar leerlo sin especificar encoding
try:
    df = pd.read_csv(csv_file)
except UnicodeDecodeError:
    # Si falla, probar con utf-16-le
    df = pd.read_csv(csv_file, encoding="utf-16-le")

print(df.head())  # Para verificar que se cargó bien

# Reemplazar valores
df = df.replace({True: "R", False: "T", "true": "R", "false": "T"})

# Guardar el DataFrame en un archivo Excel
excel_converted = "datos.xlsx"
df.to_excel(excel_converted, index=False)

# Cargar el archivo Excel para modificar estilos
wb = load_workbook(excel_converted)
ws = wb.active

# Aplicar formato a cada celda
for row in ws.iter_rows():
    for cell in row:
        if cell.value == "T":
            cell.font = Font(color="FF0000", name="Wingdings 2")  # Rojo
        elif cell.value == "R":
            cell.font = Font(color="008000", name="Wingdings 2")  # Verde

# Guardar el archivo Excel modificado
wb.save(excel_converted)

print(f"Archivo convertido y guardado como {excel_converted}")
