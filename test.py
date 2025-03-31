import openpyxl
import pandas as pd

# Nombre del archivo Excel
input_file = 'stations.xlsx'  # Cambia esto al nombre de tu archivo de entrada
output_file = 'stations_modified.xlsx'

# Cargar archivo con pandas para trabajar con los datos
df = pd.read_excel(input_file)

# Cargar el archivo con openpyxl para modificarlo
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Insertar la columna "Stations" en la primera posición (Columna A)
ws.insert_cols(1)
ws.cell(row=1, column=1).value = "Stations"  # Agregar encabezado de la nueva columna

# Variable para mantener el último valor válido
last_valid_value = None

# Iterar sobre las filas de la columna "Text" y modificar "Stations"
for row_idx, value in enumerate(df["Text"], start=2):  # Omite el encabezado y comienza en la fila 2
    if pd.isna(value):  # Si el valor en "Text" es NaN
        if last_valid_value is not None:
            ws.cell(row=row_idx, column=1).value = last_valid_value  # Propagar último valor válido
            print(last_valid_value)
    else:
        ws.cell(row=row_idx, column=1).value = value  # Usar el valor actual de "Text"
        last_valid_value = value  # Actualizar el último valor válido

# Guardar los cambios en un archivo nuevo
wb.save(output_file)
print(f"Archivo modificado guardado como {output_file}")
