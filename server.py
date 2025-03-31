from flask import Flask, request, jsonify, send_file, render_template
import pandas as pd
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os

app = Flask(__name__)

# 🔹 Habilitar CORS para todas las rutas y orígenes específicos
CORS(app, resources={r"/upload": {"origins": "*"}, r"/download/*": {"origins": "*"}})

UPLOAD_FOLDER = "uploads"
PROCESSED_FOLDER = "processed"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/upload", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return jsonify({"error": "No se ha enviado ningún archivo"}), 400

    file = request.files["file"]
    
    if file.filename == "":
        return jsonify({"error": "El nombre del archivo está vacío"}), 400

    if not file.filename.endswith(".csv"):
        return jsonify({"error": "Formato no permitido, solo CSV"}), 400

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    try:
        df = pd.read_csv(filepath)
    except UnicodeDecodeError:
        # Si falla, probar con utf-16-le
        df = pd.read_csv(filepath, encoding="utf-16-le")


    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # Reemplazar valores
    df = df.replace({True: "R", False: "T", "true": "R", "false": "T"})

    # Verificar si las columnas necesarias existen
    if "Text" not in df.columns or "TC name" not in df.columns:
        raise ValueError("Las columnas 'Text' o 'TC name' no se encuentran en el archivo CSV.")

    # Agregar columna 'Stations' en la primera posición
    df.insert(0, "Stations", None)

    # Rellenar 'Stations' con valores de 'Text' solo cuando 'TC name' sea NaN
    df.loc[df["TC name"].isna(), "Stations"] = df["Text"]

    # Propagar los valores de "Stations" hacia abajo (Forward Fill)
    df["Stations"].fillna(method="ffill", inplace=True)

    # Eliminar filas donde 'TC name' sea NaN
    df = df.dropna(subset=["TC name"]) 

    # Guardar el DataFrame en un archivo Excel
    excel_path = os.path.join(PROCESSED_FOLDER, file.filename.replace(".csv", ".xlsx"))
    df.to_excel(excel_path, index=False)
    
    # Cargar el archivo Excel para modificar estilos
    wb = load_workbook(excel_path)
    ws = wb.active

    # Definir estilo de borde
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar formato a cada celda
    for row in ws.iter_rows():
        for cell in row:
            # Aplicar bordes a todas las celdas
            cell.border = thin_border

            if cell.value == "T":
                cell.font = Font(color="FF0000", name="Wingdings 2", size=14)  # Rojo
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif cell.value == "R":
                cell.font = Font(color="008000", name="Wingdings 2", size=14)  # Verde
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Guardar el archivo Excel modificado
    wb.save(excel_path)

    response = {
        "message": "Archivo procesado exitosamente",
        "columns": df.columns.tolist(),
        "rows": df.shape[0],
        "download_link": f"/download/{file.filename.replace('.csv', '.xlsx')}"
    }
    return jsonify(response), 200


@app.route("/download/<filename>", methods=["GET"])
def download_file(filename):
    file_path = os.path.join(PROCESSED_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return jsonify({"error": "Archivo no encontrado"}), 404

if __name__ == "__main__":
    # 🔹 Importante: Escuchar en todas las interfaces y permitir conexiones externas
    app.run(debug=True, host="0.0.0.0", port=5000)